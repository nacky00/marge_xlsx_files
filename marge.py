import openpyxl
import glob
import os
import csv

def create_xlsx_file_from_csv():
    csv_files = glob.glob("./input/*.csv")
    if csv_files:
        for file in csv_files:
            wb = openpyxl.Workbook()
            name = os.path.splitext(os.path.basename(file))[0]
            wb.create_sheet(name)
            wb.active = wb.sheetnames.index(os.path.splitext(os.path.basename(file))[0])
            ws = wb.active
            with open(file, encoding="utf-8") as f:
                reader = csv.reader(f, delimiter=',')
                for row in reader:
                    ws.append(row)
            wb.save("./input/" + name + ".xlsx")
            wb.remove(wb["Sheet"])
            wb.save("./input/" + name + ".xlsx")
        wb.close()

def create_output_xlsx():
    input_files = []

    # output用ブックを作成
    output_wb = openpyxl.Workbook()
    output_wb.save("./output/output.xlsx")

    xlsx_files = glob.glob("./input/*.xlsx")

    if xlsx_files:
        for file in xlsx_files:
            # ファイル名のみ取得
            input_files.append(file)
            print(file)

        for file in input_files:
            # ブックを取得
            wb = openpyxl.load_workbook(file)
            # 先頭のシートを取得
            ws = wb.worksheets[0]
            # ペースト先のシートを取得
            paste_ws = output_wb.create_sheet(title=ws.title)
            # wsの内容をoutput_wbの新規シートにペースト
            for row in ws:
                for cell in row:
                    paste_ws[cell.coordinate].value = cell.value

        # 初期シートを削除
        output_wb.remove(output_wb["Sheet"])
        output_wb.save("./output/output.xlsx")

create_xlsx_file_from_csv()
create_output_xlsx()